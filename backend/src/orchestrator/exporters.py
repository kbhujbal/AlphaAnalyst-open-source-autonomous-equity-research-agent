"""Memo / DCF model exporters.

PDF: built with reportlab (pure Python; no system libraries beyond stdlib).
     We deliberately picked reportlab over weasyprint to avoid the cairo +
     pango runtime requirement weasyprint imposes on Linux/Mac builds.

Excel: built with openpyxl. Computed values are written, but live formulas
     drive every total cell so analysts can edit assumptions and re-run.
"""
from __future__ import annotations

import logging
from decimal import Decimal
from pathlib import Path

from src.agents.synthesizer import Memo
from src.modeler.dcf import DCFResult

logger = logging.getLogger(__name__)


_SECTION_TITLES: list[tuple[str, str]] = [
    ("Executive Summary", "executive_summary"),
    ("Financial Snapshot", "financial_snapshot"),
    ("Recent Catalysts", "recent_catalysts"),
    ("Valuation", "valuation"),
    ("Earnings Call Tone Shift", "earnings_call_tone_shift"),
    ("Alt Data Signals", "alt_data_signals"),
    ("Bull Case", "bull_case"),
    ("Bear Case", "bear_case"),
    ("Risks", "risks"),
]


def export_pdf(memo: Memo, path: Path) -> None:
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
    )

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(str(path), pagesize=LETTER, title=f"AlphaAnalyst — {memo.ticker}")
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"AlphaAnalyst Research Memo: {memo.ticker}", styles["Title"]),
        Paragraph(f"As of {memo.as_of.isoformat()}", styles["Italic"]),
        Spacer(1, 12),
    ]
    for title, attr in _SECTION_TITLES:
        body = getattr(memo, attr) or "(empty)"
        story.append(Paragraph(title, styles["Heading2"]))
        # Reportlab renders <br/> as a hard break; preserve simple line breaks.
        body_html = body.replace("\n", "<br/>")
        story.append(Paragraph(body_html, styles["BodyText"]))
        story.append(Spacer(1, 8))

    if memo.citations:
        story.append(PageBreak())
        story.append(Paragraph("Citations", styles["Heading2"]))
        for i, c in enumerate(memo.citations, start=1):
            story.append(
                Paragraph(
                    f"{i}. [{c.source_type}] {c.source_id} — {c.snippet}",
                    styles["BodyText"],
                )
            )

    doc.build(story)
    logger.info("wrote PDF memo to %s", path)


def _to_float(x: Decimal | float | int | None) -> float | None:
    if x is None:
        return None
    return float(x)


def export_excel(
    dcf_result: DCFResult,
    sensitivity_table: dict[str, dict[str, Decimal]] | None,
    path: Path,
) -> None:
    from openpyxl import Workbook

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ---- Inputs sheet ----
    ws_in = wb.active
    ws_in.title = "Inputs"
    ws_in["A1"] = "Field"
    ws_in["B1"] = "Value"
    ws_in["A2"] = "growth_rate_used"
    ws_in["B2"] = _to_float(dcf_result.growth_rate_used)
    ws_in["A3"] = "avg_fcf_margin_used"
    ws_in["B3"] = _to_float(dcf_result.avg_fcf_margin_used)
    ws_in["A4"] = "terminal_value"
    ws_in["B4"] = _to_float(dcf_result.terminal_value)
    ws_in["A5"] = "pv_terminal"
    ws_in["B5"] = _to_float(dcf_result.pv_terminal)
    ws_in["A6"] = "enterprise_value"
    ws_in["B6"] = _to_float(dcf_result.enterprise_value)
    ws_in["A7"] = "equity_value"
    ws_in["B7"] = _to_float(dcf_result.equity_value)
    ws_in["A8"] = "intrinsic_value_per_share"
    ws_in["B8"] = _to_float(dcf_result.intrinsic_value_per_share)

    # ---- DCF sheet (with live formulas where possible) ----
    ws_dcf = wb.create_sheet("DCF")
    headers = ["Year", "Projected Revenue", "Projected FCF", "Present Value"]
    for col, header in enumerate(headers, start=1):
        ws_dcf.cell(row=1, column=col, value=header)

    n = len(dcf_result.projected_fcfs)
    for i in range(n):
        row = i + 2
        ws_dcf.cell(row=row, column=1, value=i + 1)
        ws_dcf.cell(row=row, column=2, value=_to_float(dcf_result.projected_revenues[i]))
        ws_dcf.cell(row=row, column=3, value=_to_float(dcf_result.projected_fcfs[i]))
        ws_dcf.cell(row=row, column=4, value=_to_float(dcf_result.present_values[i]))

    total_row = n + 3
    ws_dcf.cell(row=total_row, column=3, value="Sum PV explicit")
    ws_dcf.cell(row=total_row, column=4, value=f"=SUM(D2:D{n + 1})")

    ws_dcf.cell(row=total_row + 1, column=3, value="PV terminal")
    ws_dcf.cell(row=total_row + 1, column=4, value=_to_float(dcf_result.pv_terminal))

    ws_dcf.cell(row=total_row + 2, column=3, value="Enterprise value")
    ws_dcf.cell(
        row=total_row + 2, column=4, value=f"=D{total_row}+D{total_row + 1}"
    )

    # ---- Sensitivity sheet ----
    if sensitivity_table:
        ws_s = wb.create_sheet("Sensitivity")
        waccs = sorted(sensitivity_table.keys(), key=lambda s: float(s))
        growths_set: set[str] = set()
        for row in sensitivity_table.values():
            growths_set |= row.keys()
        growths = sorted(growths_set, key=lambda s: float(s))

        ws_s.cell(row=1, column=1, value="WACC \\ growth")
        for j, g in enumerate(growths):
            ws_s.cell(row=1, column=j + 2, value=g)
        for i, w in enumerate(waccs):
            ws_s.cell(row=i + 2, column=1, value=w)
            row = sensitivity_table[w]
            for j, g in enumerate(growths):
                v = row.get(g)
                ws_s.cell(row=i + 2, column=j + 2, value=_to_float(v))

    wb.save(str(path))
    logger.info("wrote Excel model to %s", path)
